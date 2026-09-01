#!/usr/bin/env python3
"""豊中市の学校給食献立Excelを取得し、月別JSONへ変換する。"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
import time
import unicodedata
import warnings
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urljoin, urlsplit
from urllib.request import Request, urlopen

from openpyxl import load_workbook


BASE_URL = "https://www.city.toyonaka.osaka.jp/kosodate/gakkou/kyushoku/kondate"
USER_AGENT = "toyonaka-school-lunch/2.0 (+https://www.city.toyonaka.osaka.jp/)"
SCHEMA_VERSION = "1.0"
WEEKDAYS = "月火水木金土日"
BEVERAGE_NAMES = frozenset({"牛乳", "ぎゅうにゅう", "飲むヨーグルト", "のむヨーグルト"})
DATE_RE = re.compile(r"(?P<month>\d{1,2})月(?P<day>\d{1,2})日\((?P<weekday>[月火水木金土日])\)")
YEAR_RE = re.compile(r"\((?P<year>\d{4})年\)")
BLOCK_RE = re.compile(r"(?P<block>[A-Za-z])献立")


@dataclass(frozen=True)
class SourceConfig:
    id: str
    name: str
    level: str
    block: str | None
    url: str
    schools: tuple[str, ...]


SOURCES = (
    SourceConfig(
        "middle-a",
        "中学校 Aブロック",
        "middle_school",
        "A",
        f"{BASE_URL}/chugaku_gimukou/tyugakuR8A.html",
        ("二中", "八中", "十一中", "十三中", "十四中", "十八中"),
    ),
    SourceConfig(
        "middle-b",
        "中学校 Bブロック",
        "middle_school",
        "B",
        f"{BASE_URL}/chugaku_gimukou/tyugakuR8B.html",
        ("九中", "十二中", "十五中", "十六中", "十七中"),
    ),
    SourceConfig(
        "middle-c",
        "中学校 Cブロック",
        "middle_school",
        "C",
        f"{BASE_URL}/chugaku_gimukou/tyugakuR8C.html",
        ("一中", "三中", "四中", "五中", "庄内さくら学園（後期課程）"),
    ),
    SourceConfig(
        "middle-yotsuba",
        "庄内よつば学園（後期課程）",
        "middle_school",
        "yotsuba_late",
        f"{BASE_URL}/chugaku_gimukou/tyuugakuyotubakouki.html",
        ("庄内よつば学園（後期課程）",),
    ),
    SourceConfig(
        "elementary-a",
        "小学校 A献立",
        "elementary_school",
        "A",
        f"{BASE_URL}/shogaku_gimuzen/R8Akondate.html",
        (
            "豊島", "上野", "東丘", "刀根山", "東豊台", "箕輪", "緑地", "東泉丘",
            "庄内さくら学園（前期課程）", "豊南", "南桜塚", "北丘", "西丘", "高川",
            "豊島北", "野畑", "新田南", "蛍池", "原田", "庄内よつば学園（前期課程）",
        ),
    ),
    SourceConfig(
        "elementary-b",
        "小学校 B献立",
        "elementary_school",
        "B",
        f"{BASE_URL}/shogaku_gimuzen/R8Bkondate.html",
        (
            "桜塚", "大池", "熊野田", "中豊島", "小曽根", "南丘", "泉丘", "少路",
            "北条", "北緑丘", "克明", "桜井谷", "新田", "東豊中", "豊島西", "寺内",
            "桜井谷東",
        ),
    ),
)
SOURCE_BY_ID = {source.id: source for source in SOURCES}


class LunchDataError(RuntimeError):
    """取得データを安全に変換できない場合のエラー。"""


@dataclass(frozen=True)
class ParsedDay:
    source_id: str
    year: int
    month: int
    level: str
    block: str
    value: dict[str, object]


class _ExcelLinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.links: list[tuple[str, str]] = []
        self._href: str | None = None
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag != "a":
            return
        self._href = dict(attrs).get("href")
        self._text = []

    def handle_data(self, data: str) -> None:
        if self._href is not None:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag != "a" or self._href is None:
            return
        self.links.append((self._href, "".join(self._text).strip()))
        self._href = None
        self._text = []


def discover_workbook_urls(html: str, page_url: str) -> list[str]:
    """献立表ExcelだけをHTMLから抽出する。"""
    parser = _ExcelLinkParser()
    parser.feed(html)

    urls: list[str] = []
    seen: set[str] = set()
    for href, text in parser.links:
        if not urlsplit(href).path.lower().endswith(".xlsx"):
            continue
        if "献立表" not in text or "配合表" in text:
            continue
        url = urljoin(page_url, href)
        if url not in seen:
            seen.add(url)
            urls.append(url)
    return urls


def fetch_url_bytes(url: str, *, attempts: int = 3, timeout: float = 30.0) -> bytes:
    """一時的な通信失敗を再試行しながらURLを取得する。"""
    last_error: Exception | None = None
    request = Request(url, headers={"User-Agent": USER_AGENT})

    for attempt in range(attempts):
        try:
            with urlopen(request, timeout=timeout) as response:
                return response.read()
        except HTTPError as exc:
            if exc.code < 500:
                raise LunchDataError(f"HTTP {exc.code}: {url}") from exc
            last_error = exc
        except URLError as exc:
            last_error = exc

        if attempt + 1 < attempts:
            time.sleep(2**attempt)

    raise LunchDataError(f"取得に失敗しました: {url}: {last_error}") from last_error


def decode_html(content: bytes, content_type: str | None = None) -> str:
    """豊中市ページのHTMLを文字列へ変換する。"""
    charset_match = re.search(r"charset=([^;\s]+)", content_type or "", re.IGNORECASE)
    encodings = [charset_match.group(1).strip('"\'') if charset_match else "utf-8", "utf-8-sig", "cp932"]
    for encoding in dict.fromkeys(encodings):
        try:
            return content.decode(encoding)
        except (LookupError, UnicodeDecodeError):
            continue
    raise LunchDataError("HTMLの文字コードを判定できませんでした")


def _clean_text(value: object) -> str:
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"[\t\n]+", " ", text).strip()


def _compact_text(value: object) -> str:
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(value)))


def _worksheet_metadata(title: str, source: SourceConfig) -> tuple[int, str]:
    normalized = unicodedata.normalize("NFKC", title)
    year_match = YEAR_RE.search(normalized)
    if not year_match:
        raise LunchDataError(f"シート名から年度を判定できません: {title}")

    block = source.block
    if block is None:
        block_match = BLOCK_RE.search(normalized)
        if not block_match:
            raise LunchDataError(f"シート名からブロックを判定できません: {title}")
        block = block_match.group("block").upper()
    return int(year_match.group("year")), block


def _find_columns_and_header_row(worksheet) -> tuple[list[int], list[int], int]:
    best_row = 0
    menu_columns: list[int] = []
    for row in range(1, worksheet.max_row + 1):
        columns = [
            column
            for column in range(1, worksheet.max_column + 1)
            if _compact_text(worksheet.cell(row, column).value or "") == "献立名"
        ]
        if len(columns) > len(menu_columns):
            best_row = row
            menu_columns = columns

    if not menu_columns:
        raise LunchDataError(f"献立名の列を検出できません: {worksheet.title}")

    date_columns = [column - 1 for column in menu_columns]
    if any(column < 1 for column in date_columns):
        raise LunchDataError(f"実施日列を検出できません: {worksheet.title}")
    return date_columns, menu_columns, best_row


def _is_date_component(value: str) -> bool:
    compact = _compact_text(value)
    return (
        compact == "実施日"
        or compact in {"月", "日"}
        or bool(re.fullmatch(r"\d{1,2}", compact))
        or bool(re.fullmatch(r"\([月火水木金土日]\)", compact))
        or bool(re.fullmatch(r"〈[^〉]+〉", compact))
    )


def _extract_tags(date_parts: Iterable[str]) -> list[str]:
    tags: list[str] = []
    for value in date_parts:
        cleaned = _clean_text(value)
        if not cleaned or _is_date_component(cleaned) or cleaned in tags:
            continue
        tags.append(cleaned)
    return tags


def _is_beverage(menu_item: str) -> bool:
    return unicodedata.normalize("NFKC", menu_item) in BEVERAGE_NAMES


def parse_workbook(
    content: bytes,
    source: SourceConfig,
    *,
    source_name: str = "workbook.xlsx",
) -> list[ParsedDay]:
    """豊中市の献立Excelを日別レコードへ変換する。"""
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="DrawingML support is incomplete.*")
            workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:
        raise LunchDataError(f"Excelを開けません: {source_name}: {exc}") from exc

    parsed: list[ParsedDay] = []
    try:
        for worksheet in workbook.worksheets:
            year, block = _worksheet_metadata(worksheet.title, source)
            date_columns, menu_columns, header_row = _find_columns_and_header_row(worksheet)
            nutrition_rows = [
                row
                for row in range(header_row + 1, worksheet.max_row + 1)
                if _compact_text(worksheet.cell(row, date_columns[0]).value or "") == "栄養価"
            ]
            if not nutrition_rows:
                raise LunchDataError(f"栄養価行を検出できません: {worksheet.title}")

            region_starts = [header_row + 1] + [row + 4 for row in nutrition_rows[:-1]]
            region_ends = [row - 1 for row in nutrition_rows]

            for start_row, end_row in zip(region_starts, region_ends, strict=True):
                for date_column, menu_column in zip(date_columns, menu_columns, strict=True):
                    raw_date_parts = [
                        worksheet.cell(row, date_column).value
                        for row in range(start_row, end_row + 1)
                        if worksheet.cell(row, date_column).value not in (None, "")
                    ]
                    date_parts = [_clean_text(value) for value in raw_date_parts]
                    date_text = "".join(_compact_text(value) for value in raw_date_parts)
                    date_match = DATE_RE.search(date_text)
                    if not date_match:
                        continue

                    month = int(date_match.group("month"))
                    day_number = int(date_match.group("day"))
                    weekday = date_match.group("weekday")
                    try:
                        actual_date = date(year, month, day_number)
                    except ValueError as exc:
                        raise LunchDataError(f"不正な日付です: {worksheet.title} {date_text}") from exc
                    actual_weekday = WEEKDAYS[actual_date.weekday()]
                    if weekday != actual_weekday:
                        raise LunchDataError(
                            f"曜日が一致しません: {actual_date.isoformat()} "
                            f"Excel={weekday} 実際={actual_weekday}"
                        )

                    menu_items: list[str] = []
                    for row in range(start_row, end_row + 1):
                        value = worksheet.cell(row, menu_column).value
                        if not isinstance(value, str):
                            continue
                        item = _clean_text(value)
                        if not item or _compact_text(item) == "献立名":
                            continue
                        menu_items.append(item)

                    beverages = [item for item in menu_items if _is_beverage(item)]
                    menu = [item for item in menu_items if not _is_beverage(item)]
                    if not menu:
                        raise LunchDataError(f"献立が空です: {source.id} {actual_date.isoformat()}")

                    day_value: dict[str, object] = {
                        "date": actual_date.isoformat(),
                        "weekday": weekday,
                        "status": "scheduled",
                        "menu": menu,
                        "beverages": beverages,
                    }
                    tags = _extract_tags(date_parts)
                    if tags:
                        day_value["tags"] = tags
                    parsed.append(
                        ParsedDay(source.id, year, month, source.level, block, day_value)
                    )
    finally:
        workbook.close()

    if not parsed:
        raise LunchDataError(f"献立を1件も抽出できません: {source.id} {source_name}")
    return parsed


DocumentKey = tuple[str, int, int]


def build_month_documents(records: Iterable[ParsedDay]) -> dict[DocumentKey, dict[str, object]]:
    """日別レコードを取得元・月単位にまとめ、矛盾する重複を拒否する。"""
    grouped: dict[DocumentKey, list[ParsedDay]] = defaultdict(list)
    for record in records:
        grouped[(record.source_id, record.year, record.month)].append(record)

    documents: dict[DocumentKey, dict[str, object]] = {}
    for key, month_records in grouped.items():
        levels = {record.level for record in month_records}
        blocks = {record.block for record in month_records}
        if len(levels) != 1 or len(blocks) != 1:
            raise LunchDataError(f"同じ出力月の学校区分またはブロックが一致しません: {key}")

        days_by_date: dict[str, dict[str, object]] = {}
        for record in month_records:
            date_string = str(record.value["date"])
            previous = days_by_date.get(date_string)
            if previous is not None and previous != record.value:
                raise LunchDataError(f"同じ日付に異なる献立があります: {key[0]} {date_string}")
            days_by_date[date_string] = record.value

        _, year, month = key
        documents[key] = {
            "schema_version": SCHEMA_VERSION,
            "year": year,
            "month": month,
            "level": next(iter(levels)),
            "block": next(iter(blocks)),
            "days": [days_by_date[value] for value in sorted(days_by_date)],
        }
    return documents


def _atomic_write_text(destination: Path, content: str) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            newline="\n",
            dir=destination.parent,
            prefix=f".{destination.name}.",
            suffix=".tmp",
            delete=False,
        ) as temporary:
            temporary.write(content)
            temporary_path = Path(temporary.name)
        os.replace(temporary_path, destination)
    finally:
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()


def write_documents(documents: dict[DocumentKey, dict[str, object]], output_dir: Path) -> list[Path]:
    """すべてのJSONを直列化してから、ファイル単位で原子的に保存する。"""
    serialized = {
        key: json.dumps(document, ensure_ascii=False, indent=2) + "\n"
        for key, document in documents.items()
    }

    written: list[Path] = []
    for (source_id, year, month), content in sorted(serialized.items()):
        destination = output_dir / source_id / f"{year:04d}-{month:02d}.json"
        _atomic_write_text(destination, content)
        written.append(destination)
    return written


def write_source_index(output_dir: Path, source_configs: Iterable[SourceConfig]) -> Path:
    """既存の月別ファイルを走査し、Web/API用の決定的な索引を生成する。"""
    entries: list[dict[str, object]] = []
    for source in source_configs:
        months = sorted(path.stem for path in (output_dir / source.id).glob("????-??.json"))
        if not months:
            continue
        entries.append(
            {
                "id": source.id,
                "name": source.name,
                "level": source.level,
                "block": source.block or "auto",
                "schools": list(source.schools),
                "source_url": source.url,
                "months": months,
            }
        )

    index = {"schema_version": SCHEMA_VERSION, "sources": entries}
    destination = output_dir / "index.json"
    _atomic_write_text(destination, json.dumps(index, ensure_ascii=False, indent=2) + "\n")
    return destination


def run(source_configs: Iterable[SourceConfig], output_dir: Path) -> list[Path]:
    selected_sources = tuple(source_configs)
    all_records: list[ParsedDay] = []
    for source in selected_sources:
        html = decode_html(fetch_url_bytes(source.url))
        workbook_urls = discover_workbook_urls(html, source.url)
        if not workbook_urls:
            raise LunchDataError(f"献立表Excelのリンクが見つかりませんでした: {source.id}")
        for workbook_url in workbook_urls:
            content = fetch_url_bytes(workbook_url)
            source_name = Path(urlsplit(workbook_url).path).name
            all_records.extend(parse_workbook(content, source, source_name=source_name))

    documents = build_month_documents(all_records)
    written = write_documents(documents, output_dir)
    known_sources = {source.id: source for source in SOURCES}
    known_sources.update({source.id: source for source in selected_sources})
    written.append(write_source_index(output_dir, known_sources.values()))
    return written


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    selection = parser.add_mutually_exclusive_group()
    selection.add_argument(
        "--source",
        action="append",
        choices=sorted(SOURCE_BY_ID),
        help="更新する取得元。複数指定可能（省略時は全取得元）",
    )
    selection.add_argument(
        "--url",
        help="任意の献立表ページ。登録済みURLなら対応する取得元として処理",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data"),
        help="月別JSONの出力先（既定: data）",
    )
    return parser.parse_args(argv)


def _select_sources(args: argparse.Namespace) -> tuple[SourceConfig, ...]:
    if args.url:
        normalized_url = args.url.rstrip("/")
        known = next((source for source in SOURCES if source.url.rstrip("/") == normalized_url), None)
        if known:
            return (known,)
        return (SourceConfig("custom", "カスタム", "other", None, args.url, ()),)
    if args.source:
        return tuple(SOURCE_BY_ID[source_id] for source_id in args.source)
    return SOURCES


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    selected_sources = _select_sources(args)
    try:
        written = run(selected_sources, args.output_dir)
    except LunchDataError as exc:
        print(f"エラー: {exc}", file=sys.stderr)
        return 1

    for path in written:
        if path.name == "index.json":
            print(f"{path}: 取得元索引")
            continue
        with path.open(encoding="utf-8") as file:
            day_count = len(json.load(file)["days"])
        print(f"{path}: {day_count}日分")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
